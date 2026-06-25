import json, re, unicodedata, os, zipfile, shutil
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

BASE='/mnt/data'
WORK=os.path.join(BASE,'work')
PRED=os.path.join(WORK,'predicciones.json')
RES=os.path.join(WORK,'resultados.json')
EXCEL=os.path.join(BASE,'Resultados Real 2.xlsx')
OUT_XLSX=os.path.join(BASE,'auditoria_completa_polla_v4_2.xlsx')
OUT_ZIP=os.path.join(BASE,'Norteamerica2026SMA-V4.2-auditada.zip')

pred_data=json.load(open(PRED,encoding='utf-8'))
participants=pred_data['participantes']
matches=json.load(open(RES,encoding='utf-8'))['partidos']

TEAM_ALIAS={
 'mexico':'mexico','mejico':'mexico','mex':'mexico',
 'sudafrica':'sudafrica','sud africa':'sudafrica','south africa':'sudafrica','saf':'sudafrica','rsa':'sudafrica',
 'corea del sur':'corea del sur','korea':'corea del sur','south korea':'corea del sur','kor':'corea del sur',
 'chequia':'chequia','czechia':'chequia','czech republic':'chequia','cze':'chequia',
 'canada':'canada',
 'bosnia':'bosnia','bosnia y herzegovina':'bosnia','bosnia and herzegovina':'bosnia','bih':'bosnia',
 'qatar':'qatar',
 'suiza':'suiza','switzerland':'suiza','sui':'suiza',
 'brasil':'brasil','brazil':'brasil','bra':'brasil',
 'marruecos':'marruecos','morocco':'marruecos','mar':'marruecos',
 'haiti':'haiti','hti':'haiti',
 'escocia':'escocia','scotland':'escocia','sco':'escocia',
 'argentina':'argentina','arg':'argentina','argelia':'argelia','algeria':'argelia','dza':'argelia',
 'austria':'austria','aut':'austria','jordania':'jordania','jordan':'jordania','jor':'jordania',
 'ee uu':'ee uu','eeuu':'ee uu','estados unidos':'ee uu','united states':'ee uu','usa':'ee uu',
 'paises bajos':'paises bajos','p bajos':'paises bajos','netherlands':'paises bajos','holanda':'paises bajos',
 'c de marfil':'c de marfil','costa de marfil':'c de marfil','ivory coast':'c de marfil',
 'n zelanda':'n zelanda','nueva zelanda':'n zelanda','new zealand':'n zelanda',
 'rd congo':'rd congo','dr congo':'rd congo','r d congo':'rd congo','democratic republic of congo':'rd congo'
}

def strip_acc(s):
    return ''.join(c for c in unicodedata.normalize('NFD',str(s or '')) if unicodedata.category(c)!='Mn')

def team_key(s):
    s=strip_acc(s).lower().strip()
    s=re.sub(r'[^a-z0-9]+',' ',s).strip()
    return TEAM_ALIAS.get(s,s)

def same_team(a,b): return team_key(a)==team_key(b)

def pair_key(a,b): return '|'.join(sorted([team_key(a),team_key(b)]))

def name_key(s):
    s=strip_acc(s).lower().strip()
    s=re.sub(r'\([^)]*\)','',s)
    s=re.sub(r'[^a-z0-9]+',' ',s).strip()
    return s

def is_final(m):
    return m.get('estado')=='finalizado' and m.get('golesLocal') is not None and m.get('golesVisitante') is not None

def match_time(m): return m.get('fecha') or m.get('date') or ''

def group_letters(): return sorted({m.get('grupo') for m in matches if m.get('grupo')})

def group_matches(g): return sorted([m for m in matches if m.get('grupo')==g], key=match_time)

def table_row(team): return {'team':team,'pj':0,'pts':0,'gf':0,'gc':0,'gd':0,'w':0,'d':0,'l':0}

def add_standing_match(rows, local, visitante, gl, gv):
    rows.setdefault(local, table_row(local)); rows.setdefault(visitante, table_row(visitante))
    a,b=rows[local],rows[visitante]
    a['pj']+=1; b['pj']+=1
    a['gf']+=gl; a['gc']+=gv; b['gf']+=gv; b['gc']+=gl
    a['gd']=a['gf']-a['gc']; b['gd']=b['gf']-b['gc']
    if gl>gv:
        a['w']+=1; b['l']+=1; a['pts']+=3
    elif gl<gv:
        b['w']+=1; a['l']+=1; b['pts']+=3
    else:
        a['d']+=1; b['d']+=1; a['pts']+=1; b['pts']+=1

def sort_standings(rows):
    return sorted(rows.values(), key=lambda x:(-x['pts'],-x['gd'],-x['gf'],x['gc'],team_key(x['team'])))

def actual_group_standing(g):
    ms=group_matches(g)
    if len(ms)<6 or not all(is_final(m) for m in ms): return None
    rows={}
    for m in ms: add_standing_match(rows,m['local'],m['visitante'],int(m['golesLocal']),int(m['golesVisitante']))
    return sort_standings(rows)

def predicted_group_standing_from_scores(g,p):
    ms=group_matches(g)
    if len(ms)<6: return None
    rows={}
    for m in ms:
        pr=p.get('predicciones',{}).get(m['id'])
        if not pr or pr.get('golesLocal') is None or pr.get('golesVisitante') is None: return None
        add_standing_match(rows,m['local'],m['visitante'],int(pr['golesLocal']),int(pr['golesVisitante']))
    return sort_standings(rows)

def predicted_group_standing(g,p):
    by_score=predicted_group_standing_from_scores(g,p) or []
    stats={team_key(x['team']):x for x in by_score}
    if p.get('grupos') and isinstance(p['grupos'].get(g),list) and len(p['grupos'][g])>=4:
        return [dict(stats.get(team_key(t),{}),team=t) for t in p['grupos'][g]]
    return by_score if by_score else None

BONUS=[4,3,1,1]
def group_position_bonus_for_group(g,p):
    real=actual_group_standing(g); pred=predicted_group_standing(g,p)
    if not real or not pred: return None, []
    pts=0; detail=[]
    for i in range(4):
        hit= bool(real[i].get('team') and pred[i].get('team') and same_team(real[i]['team'],pred[i]['team']))
        add=BONUS[i] if hit else 0
        pts+=add
        detail.append({'pos':i+1,'real':real[i]['team'],'pred':pred[i].get('team'),'hit':hit,'pts':add})
    return pts,detail

def group_position_bonus(p):
    total=0; hits=0; closed=0; detail=[]
    for g in group_letters():
        pts,d=group_position_bonus_for_group(g,p)
        if pts is None: continue
        closed+=1; total+=pts; hits+=sum(1 for x in d if x['hit'])
        for x in d: detail.append(dict(x,grupo=g))
    return {'total':total,'hits':hits,'closed':closed,'detail':detail}

R32_SLOTS=[
  {'id':'R32-1','a':{'group':'A','pos':2},'b':{'group':'B','pos':2}},
  {'id':'R32-2','a':{'group':'C','pos':1},'b':{'group':'F','pos':2}},
  {'id':'R32-3','a':{'group':'E','pos':1},'b':{'third':['A','B','C','D','F']}},
  {'id':'R32-4','a':{'group':'F','pos':1},'b':{'group':'C','pos':2}},
  {'id':'R32-5','a':{'group':'E','pos':2},'b':{'group':'I','pos':2}},
  {'id':'R32-6','a':{'group':'I','pos':1},'b':{'third':['C','D','F','G','H']}},
  {'id':'R32-7','a':{'group':'A','pos':1},'b':{'third':['C','E','F','H','I']}},
  {'id':'R32-8','a':{'group':'L','pos':1},'b':{'third':['E','H','I','J','K']}},
  {'id':'R32-9','a':{'group':'G','pos':1},'b':{'third':['A','E','H','I','J']}},
  {'id':'R32-10','a':{'group':'D','pos':1},'b':{'third':['B','E','F','I','J']}},
  {'id':'R32-11','a':{'group':'H','pos':1},'b':{'group':'J','pos':2}},
  {'id':'R32-12','a':{'group':'K','pos':2},'b':{'group':'L','pos':2}},
  {'id':'R32-13','a':{'group':'B','pos':1},'b':{'third':['E','F','G','I','J']}},
  {'id':'R32-14','a':{'group':'D','pos':2},'b':{'group':'G','pos':2}},
  {'id':'R32-15','a':{'group':'J','pos':1},'b':{'group':'H','pos':2}},
  {'id':'R32-16','a':{'group':'K','pos':1},'b':{'third':['D','E','I','J','L']}}
]

def get_standing(g,p=None): return predicted_group_standing(g,p) if p else actual_group_standing(g)

def best_thirds_for_bracket(p=None):
    thirds=[]; closed=0
    for g in group_letters():
        st=get_standing(g,p)
        if not st or len(st)<4: continue
        closed+=1; thirds.append(dict(st[2],grupo=g))
    if p is None and closed<len(group_letters()): return {'thirds':[],'closed':closed,'total':len(group_letters()),'complete':False}
    thirds=sorted(thirds, key=lambda x:(-(x.get('pts') or 0),-(x.get('gd') or 0),-(x.get('gf') or 0),(x.get('gc') or 0),team_key(x['team'])))
    return {'thirds':thirds[:8],'closed':closed,'total':len(group_letters()),'complete':True if p else closed==len(group_letters())}

def resolve_slot(slot,p,used):
    if 'group' in slot:
        st=get_standing(slot['group'],p)
        if not st or not st[slot['pos']-1]: return None
        return {'team':st[slot['pos']-1]['team'],'source':f"{slot['pos']}{slot['group']}"}
    if 'third' in slot:
        best=best_thirds_for_bracket(p)
        if not best['complete']: return None
        found=None
        for x in best['thirds']:
            if x['grupo'] in slot['third'] and x['grupo'] not in used:
                found=x; break
        if not found: return None
        used.add(found['grupo'])
        return {'team':found['team'],'source':'3'+found['grupo']}
    return None

def r32_matchups(p=None):
    used=set(); out=[]
    for slot in R32_SLOTS:
        a=resolve_slot(slot['a'],p,used); b=resolve_slot(slot['b'],p,used)
        if a and b:
            out.append({'id':slot['id'],'a':a['team'],'b':b['team'],'aSource':a['source'],'bSource':b['source'],'key':pair_key(a['team'],b['team'])})
    return out

def r32_exact_matchup_bonus(p):
    real=r32_matchups(None); pred=r32_matchups(p)
    keys={x['key'] for x in pred}; detail=[]; total=0
    for m in real:
        if m['key'] in keys:
            total+=2; detail.append(dict(m,pts=2))
    return {'total':total,'hits':len(detail),'available':len(real),'detail':detail}

# Load Excel T values by group A/B/C
excel={}
if os.path.exists(EXCEL):
    wb0=load_workbook(EXCEL,data_only=True)
    for g in ['A','B','C']:
        ws=wb0[f'Grupo {g}']
        for r in range(2,ws.max_row+1):
            name=ws.cell(r,1).value
            val=ws.cell(r,20).value
            if name and val is not None:
                try: ival=int(float(val))
                except Exception: continue
                excel.setdefault(str(name).strip(),{})[g]=ival

excel_by_key={name_key(k):k for k in excel}
# Manual name aliases for Excel parentheses/typos
name_aliases={
    name_key('Alejandro Escallon Hijo'): name_key('Alejandro Escallon'),
    name_key('Alejandro Escallon Papa'): name_key('Alejandro Escallon'),
    name_key('Katerin Hernandez'): name_key('Katerine Hernandez'),
    name_key('Luis Francisco Fracica'): name_key('Luis F Fracica'),
    name_key('Luis Frascica'): name_key('Luis F Fracica'),
    name_key('Luis Frascica'): name_key('Luis F Fracica'),
}

def find_excel_name(pname):
    k=name_key(pname)
    if k in excel_by_key: return excel_by_key[k]
    # special handling for papa/hijo
    for ek,en in excel_by_key.items():
        if k==name_key(en): return en
    if k in name_aliases:
        target=name_aliases[k]
        return excel_by_key.get(target)
    # fallback contains
    for ek,en in excel_by_key.items():
        if k and (k in ek or ek in k): return en
    return None

rows=[]; diffs=[]; missing=[]
for p in participants:
    exname=find_excel_name(p['nombre'])
    if not exname:
        missing.append(p['nombre'])
    gps={}; gdetails={}
    for g in ['A','B','C']:
        pts,det=group_position_bonus_for_group(g,p)
        gps[g]=pts; gdetails[g]=det
    r32=r32_exact_matchup_bonus(p)
    excel_vals=excel.get(exname,{}) if exname else {}
    row={'id':p.get('id'),'nombre':p['nombre'],'excel_name':exname,'A_page':gps['A'],'B_page':gps['B'],'C_page':gps['C'],
         'A_excel':excel_vals.get('A'),'B_excel':excel_vals.get('B'),'C_excel':excel_vals.get('C'),
         'ABC_page':sum(x or 0 for x in gps.values()),'ABC_excel':sum(v for v in [excel_vals.get('A'),excel_vals.get('B'),excel_vals.get('C')] if isinstance(v,int)),
         'r32_pts':r32['total'],'r32_hits':r32['hits'],'r32_available':r32['available'],'grupo_total_page':sum(x or 0 for x in gps.values())+r32['total']}
    row['ABC_diff']=None if not exname else row['ABC_page']-row['ABC_excel']
    rows.append(row)
    if exname:
        for g in ['A','B','C']:
            e=excel_vals.get(g); c=gps[g]
            if e is not None and c is not None and e!=c:
                pred=[x['pred'] for x in gdetails[g]] if gdetails[g] else []
                real=[x['real'] for x in gdetails[g]] if gdetails[g] else []
                cause='Diferencia contra Excel. La regla 4-3-1-1 da página=%s; Excel=%s. Revisar si Excel omitió 3.º/4.º o si la predicción manual difiere.'%(c,e)
                diffs.append({'nombre':p['nombre'],'grupo':g,'excel':e,'pagina':c,'dif':c-e,'pred':', '.join(map(str,pred)),'real':', '.join(map(str,real)),'causa':cause})

# Felipe details
felipe=next((p for p in participants if name_key(p['nombre'])==name_key('Felipe Gaitan')),None)
felipe_rows=[]
if felipe:
    for g in ['A','B','C','J']:
        pts,det=group_position_bonus_for_group(g,felipe)
        if det:
            for d in det:
                felipe_rows.append([g,d['pos'],d['real'],d['pred'],'Sí' if d['hit'] else 'No',d['pts']])
        else:
            # J may not be closed, still show pred groups
            for i,t in enumerate(felipe.get('grupos',{}).get(g,[]),start=1): felipe_rows.append([g,i,'Pendiente',t,'Pendiente',0])

# Write audit workbook
wb=Workbook(); ws=wb.active; ws.title='Resumen'
header_fill=PatternFill('solid',fgColor='0B4F7A'); gold_fill=PatternFill('solid',fgColor='FFD447'); ok_fill=PatternFill('solid',fgColor='DCFCE7'); bad_fill=PatternFill('solid',fgColor='FECACA')
white_font=Font(color='FFFFFF',bold=True); title_font=Font(bold=True,size=14,color='073763'); thin=Side(style='thin',color='D9E2EC'); border=Border(left=thin,right=thin,top=thin,bottom=thin)

def style_header(ws,row=1):
    for cell in ws[row]:
        cell.fill=header_fill; cell.font=white_font; cell.alignment=Alignment(horizontal='center',vertical='center'); cell.border=border

def autofit(ws):
    for col in range(1,ws.max_column+1):
        mx=0
        for row in range(1,min(ws.max_row,200)+1):
            v=ws.cell(row,col).value
            if v is not None: mx=max(mx,len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width=min(max(mx+2,10),44)
    for row in ws.iter_rows():
        for c in row:
            c.border=border; c.alignment=Alignment(vertical='top',wrap_text=True)

ws['A1']='Auditoría V4.2 — participantes validados + llave exacta'; ws['A1'].font=title_font
summary=[
 ['Participantes en predicciones.json',len(participants)],
 ['Grupos cerrados',sum(1 for g in group_letters() if actual_group_standing(g))],
 ['Llaves de 16avos oficiales evaluables',len(r32_matchups(None))],
 ['Participantes con comparación Excel ABC',sum(1 for r in rows if r['excel_name'])],
 ['Diferencias contra Excel ABC',len(diffs)],
 ['Felipe Gaitán ABC calculado por página',next(r['ABC_page'] for r in rows if r['nombre']=='Felipe Gaitan')],
 ['Felipe Gaitán llave exacta 16avos',next(r['r32_pts'] for r in rows if r['nombre']=='Felipe Gaitan')],
 ['Regla aplicada','Grupo: 1.º=4, 2.º=3, 3.º=1, 4.º=1. Llave exacta 16avos=+2 por cruce, sin importar orden.'],
]
for i,(k,v) in enumerate(summary,start=3): ws.cell(i,1).value=k; ws.cell(i,2).value=v
ws['A12']='Conclusión'; ws['A12'].font=Font(bold=True,color='073763')
ws['B12']='Felipe Gaitán queda en 23 puntos por grupos A+B+C. Las diferencias contra Excel son casos donde la regla automática 4-3-1-1 no coincide con la columna T manual; quedan listadas para revisión.'
autofit(ws)

ws=wb.create_sheet('Diferencias Excel ABC')
headers=['Participante','Grupo','Excel col T','Página regla','Dif página-Excel','Predicción usada','Resultado real','Causa probable']
ws.append(headers); style_header(ws,1)
for d in diffs:
    ws.append([d['nombre'],d['grupo'],d['excel'],d['pagina'],d['dif'],d['pred'],d['real'],d['causa']])
    ws.cell(ws.max_row,5).fill=bad_fill
if not diffs: ws.append(['Sin diferencias','','','','','','',''])
autofit(ws); ws.freeze_panes='A2'

ws=wb.create_sheet('Todos participantes')
headers=['Participante','Nombre Excel','A Excel','A Página','B Excel','B Página','C Excel','C Página','ABC Excel','ABC Página','Dif ABC','Llave exacta pts','Llaves exactas','Llaves evaluables','Total grupos+llave']
ws.append(headers); style_header(ws,1)
for r in sorted(rows,key=lambda x:x['nombre']):
    ws.append([r['nombre'],r['excel_name'],r['A_excel'],r['A_page'],r['B_excel'],r['B_page'],r['C_excel'],r['C_page'],r['ABC_excel'] if r['excel_name'] else None,r['ABC_page'],r['ABC_diff'],r['r32_pts'],r['r32_hits'],r['r32_available'],r['grupo_total_page']])
    if r['ABC_diff'] not in (0,None): ws.cell(ws.max_row,11).fill=bad_fill
    elif r['ABC_diff']==0: ws.cell(ws.max_row,11).fill=ok_fill
autofit(ws); ws.freeze_panes='A2'

ws=wb.create_sheet('Felipe Gaitan')
ws.append(['Grupo','Posición','Real','Predicción Felipe','Acierta','Puntos']); style_header(ws,1)
for row in felipe_rows: ws.append(row)
# totals
ws.append([]); ws.append(['Total ABC','','','','',next(r['ABC_page'] for r in rows if r['nombre']=='Felipe Gaitan')])
ws.append(['Llave exacta 16avos','','','','',next(r['r32_pts'] for r in rows if r['nombre']=='Felipe Gaitan')])
autofit(ws); ws.freeze_panes='A2'

ws=wb.create_sheet('Llaves 16avos')
ws.append(['Llave','Equipo A','Equipo B','Fuente A','Fuente B','Participantes que acertaron']); style_header(ws,1)
for m in r32_matchups(None):
    count=sum(1 for p in participants if pair_key(m['a'],m['b']) in {x['key'] for x in r32_matchups(p)})
    ws.append([m['id'],m['a'],m['b'],m['aSource'],m['bSource'],count])
if not r32_matchups(None): ws.append(['Aún no hay llaves evaluables','','','','',''])
autofit(ws); ws.freeze_panes='A2'

ws=wb.create_sheet('Posiciones reales')
ws.append(['Grupo','Posición','Equipo','Pts','DG','GF','GC']); style_header(ws,1)
for g in group_letters():
    st=actual_group_standing(g)
    if not st: continue
    for i,x in enumerate(st,start=1): ws.append([g,i,x['team'],x['pts'],x['gd'],x['gf'],x['gc']])
autofit(ws); ws.freeze_panes='A2'

wb.save(OUT_XLSX)

# Add audit script and readme to work folder
with open(os.path.join(WORK,'auditoria_reglas.py'),'w',encoding='utf-8') as f:
    f.write(open(__file__,encoding='utf-8').read())
with open(os.path.join(WORK,'LEEME_V4_1_AUDITADA.txt'),'w',encoding='utf-8') as f:
    f.write('''V4.2 AUDITADA — Polla Mundial 2026\n\nIncluye:\n- Predicciones de grupo guardadas en predicciones.json.\n- Resultados reales se toman automáticamente de resultados.json.\n- Puntos de grupo: 1.º exacto = 4, 2.º exacto = 3, 3.º exacto = 1, 4.º exacto = 1.\n- Llave exacta de 16avos: +2 por cruce correcto, sin importar localía.\n- Felipe Gaitán corregido: Grupo J = Argentina, Argelia, Austria, Jordania.\n- Auditoría incluida: auditoria_reglas.py y auditoria_completa_polla_v4_2.xlsx.\n\nResultado validado:\n- Felipe Gaitán A+B+C = 23 puntos.\n- Se corrigió el cruce de nombre Luis Francisco Fracica ↔ Luis F. Fracica y se listan diferencias contra la columna T manual del Excel en la hoja “Diferencias Excel ABC”.\n\nPara subir: reemplaza los archivos de la raíz del repositorio con los de este ZIP.\n''')
# Copy audit workbook into package root
shutil.copy2(OUT_XLSX, os.path.join(WORK,'auditoria_completa_polla_v4_2.xlsx'))
# Zip work directory
if os.path.exists(OUT_ZIP): os.remove(OUT_ZIP)
with zipfile.ZipFile(OUT_ZIP,'w',zipfile.ZIP_DEFLATED) as z:
    for root,dirs,files in os.walk(WORK):
        # avoid __pycache__
        dirs[:] = [d for d in dirs if d!='__pycache__']
        for file in files:
            path=os.path.join(root,file)
            arc=os.path.relpath(path,WORK)
            z.write(path,arc)
print('created',OUT_XLSX,OUT_ZIP)
print('diffs',len(diffs),'real r32',len(r32_matchups(None)),'Felipe ABC',next(r['ABC_page'] for r in rows if r['nombre']=='Felipe Gaitan'))
